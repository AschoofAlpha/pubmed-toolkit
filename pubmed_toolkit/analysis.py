"""Reusable analysis pipeline for downloader outputs."""

from __future__ import annotations

import csv
import glob
import json
import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

EXCEL_COLUMNS = {
    "PMID": "pmid",
    "标题": "title",
    "作者": "authors_str",
    "角色": "role",
    "期刊": "journal",
    "发表日期": "pub_date",
    "卷": "volume",
    "期": "issue",
    "页码": "pages",
    "DOI": "doi",
    "PMC ID": "pmc_id",
    "PDF状态": "pdf_status",
    "摘要": "abstract",
}

BUCKET_NAMES = {
    "A": "A 经典通路机制\n(signaling/泛素化/cell-death)",
    "B": "B 非编码 RNA / 表观遗传",
    "C": "C 纳米医学与递送",
    "D": "D 免疫治疗 / TME / cGAS-STING",
    "E": "E 临床研究 (cohort/RCT)",
    "F": "F 其他 (生信/文献为主)",
}

BUCKET_COLOR = {
    "A": "#1f77b4",
    "B": "#9467bd",
    "C": "#2ca02c",
    "D": "#d62728",
    "E": "#bcbd22",
    "F": "#7f7f7f",
}


def find_latest_excel(output_dir: str) -> str:
    files = glob.glob(os.path.join(output_dir, "papers_*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No papers_*.xlsx found in {output_dir}")
    return max(files, key=os.path.getmtime)


def find_latest_json(output_dir: str) -> str | None:
    files = glob.glob(os.path.join(output_dir, "papers_*.json"))
    return max(files, key=os.path.getmtime) if files else None


def read_papers_from_excel(excel_path: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    papers = []
    for row in rows[1:]:
        rec = {}
        for idx, value in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            key = EXCEL_COLUMNS.get(header)
            if key:
                rec[key] = "" if value is None else str(value)
        if rec.get("pmid"):
            papers.append(rec)
    wb.close()
    return papers


def load_papers_json(path: str | None) -> list[dict]:
    if not path or not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def find_pdf(pdf_dir: str, pmid: str) -> str | None:
    cands = [
        path for path in glob.glob(os.path.join(pdf_dir, f"{pmid}_*.pdf"))
        if f"{os.sep}suspect{os.sep}" not in path
    ]
    return cands[0] if cands else None


def extract_full_text(path: str) -> str:
    try:
        import fitz  # type: ignore
    except ImportError:
        return "__ERROR__:PyMuPDF unavailable"

    try:
        doc = fitz.open(path)
        try:
            pages = [page.get_text("text") for page in doc]
        finally:
            doc.close()
        return "\n".join(pages)
    except Exception as e:
        return f"__ERROR__:{type(e).__name__}: {e}"


def split_by_section(full_text: str) -> dict[str, str]:
    if full_text.startswith("__ERROR__"):
        return {"_error": full_text}
    headings = [
        "abstract", "introduction", "background", "results", "methods",
        "materials and methods", "methodology", "discussion", "conclusion",
        "conclusions", "references", "acknowledgements", "acknowledgments",
    ]
    pattern = re.compile(
        r"(?im)^\s*(" + "|".join(re.escape(h) for h in headings) + r")\s*\.?\s*$",
        re.MULTILINE,
    )
    matches = list(pattern.finditer(full_text))
    sections = {}
    for idx, match in enumerate(matches):
        name = match.group(1).strip().lower()
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(full_text)
        sections.setdefault(name, full_text[start:end].strip())
    return sections


def _first_paragraph(text: str, max_chars: int = 1500) -> str:
    paras = [p.strip() for p in re.split(r"\n\s*\n", text or "") if p.strip()]
    return paras[0][:max_chars] if paras else ""


def _last_paragraph(text: str, max_chars: int = 1500) -> str:
    paras = [p.strip() for p in re.split(r"\n\s*\n", text or "") if p.strip()]
    return paras[-1][:max_chars] if paras else ""


def build_corpus(papers: list[dict], output_dir: str, pdf_dir: str) -> str:
    corpus = []
    for paper in papers:
        pmid = str(paper.get("pmid", ""))
        rec = {
            "pmid": pmid,
            "title": paper.get("title", ""),
            "pub_date": paper.get("pub_date", ""),
            "journal": paper.get("journal", ""),
            "abstract": paper.get("abstract", ""),
            "pdf_path": find_pdf(pdf_dir, pmid),
            "pdf_status": "missing",
            "intro_last": "",
            "disc_first": "",
            "extraction_note": "",
        }
        if rec["pdf_path"]:
            full = extract_full_text(rec["pdf_path"])
            if full.startswith("__ERROR__"):
                rec["pdf_status"] = "error"
                rec["extraction_note"] = full
            else:
                rec["pdf_status"] = "ok"
                rec["pdf_chars"] = len(full)
                sections = split_by_section(full)
                intro = sections.get("introduction") or sections.get("background") or ""
                discussion = sections.get("discussion") or ""
                rec["intro_last"] = _last_paragraph(intro)
                rec["disc_first"] = _first_paragraph(discussion)
                rec["sections_found"] = list(sections.keys())
                if not intro and not discussion:
                    rec["extraction_note"] = "no Introduction/Discussion headers parsed; abstract-only"
        else:
            rec["extraction_note"] = "no PDF available; abstract-only"
        corpus.append(rec)

    out = os.path.join(output_dir, "_paper_corpus.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(corpus, f, ensure_ascii=False, indent=2)
    return out


def _date_iso(value: str) -> str:
    value = str(value or "").strip()
    match = re.search(r"(\d{4})", value)
    year = int(match.group(1)) if match else 1900
    month = 1
    day = 1
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    month_match = re.search(r"\b([A-Za-z]{3})[A-Za-z]*\b", value)
    if month_match:
        month = month_map.get(month_match.group(1).lower(), 1)
    nums = [int(n) for n in re.findall(r"\b\d{1,2}\b", value)]
    if nums:
        day = max(1, min(nums[-1], 31))
    return f"{year:04d}-{month:02d}-{day:02d}"


def _split_author_names(authors_str: str) -> list[str]:
    return [part.strip() for part in str(authors_str or "").split(",") if part.strip()]


def build_author_records(papers: list[dict], json_papers: list[dict], output_dir: str) -> str:
    detailed_by_pmid = {str(p.get("pmid", "")): p for p in json_papers if p.get("authors")}
    records = []
    for paper in papers:
        pmid = str(paper.get("pmid", ""))
        detailed = detailed_by_pmid.get(pmid, {})
        authors = detailed.get("authors") or [
            {"name": name, "equal_contrib": False, "is_corresponding": False}
            for name in _split_author_names(paper.get("authors_str", ""))
        ]
        parsed_authors = []
        for idx, author in enumerate(authors):
            name = author.get("name", "")
            if not name:
                continue
            parsed_authors.append({
                "name": name,
                "index": idx + 1,
                "is_first": idx == 0,
                "is_last": idx == len(authors) - 1,
                "equal_contrib": bool(author.get("equal_contrib")),
                "is_corresponding": bool(author.get("is_corresponding")),
                "affiliation": author.get("affiliation", ""),
            })
        records.append({
            "pmid": pmid,
            "title": paper.get("title", ""),
            "pubdate": _date_iso(paper.get("pub_date", "")),
            "authors": parsed_authors,
        })

    out = os.path.join(output_dir, "_authors_parsed.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    return out


def _author_role_label(author: dict) -> str:
    labels = []
    if author.get("is_first"):
        labels.append("F")
    if author.get("equal_contrib"):
        labels.append("cF")
    if author.get("is_last"):
        labels.append("L")
    if author.get("is_corresponding"):
        labels.append("C")
    if not labels:
        labels.append(f"#{author.get('index', '')}")
    return " / ".join(labels)


def write_author_matrix(author_records_path: str, output_dir: str) -> str:
    with open(author_records_path, encoding="utf-8") as f:
        records = json.load(f)
    records.sort(key=lambda r: r.get("pubdate", ""))

    authors = sorted({
        author["name"]
        for rec in records
        for author in rec.get("authors", [])
        if author.get("name")
    })
    columns = [f"{rec['pmid']} ({rec.get('pubdate', '')})" for rec in records]
    role_map = defaultdict(dict)
    # strict=True asserts the invariant that `columns` was built from `records`.
    for rec, col in zip(records, columns, strict=True):
        for author in rec.get("authors", []):
            role_map[author["name"]][col] = _author_role_label(author)

    out = os.path.join(output_dir, "_author_paper_matrix.csv")
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["author"] + columns)
        for author in authors:
            writer.writerow([author] + [role_map[author].get(col, "") for col in columns])
    return out


# --- Activity timeline (docs/profile-visual-spec.md Section 4) ---
#
# Geometry is fixed in pixels rather than left to matplotlib's autoscaling. The
# measured failure this replaces was 1934x21506 px — a 1:11 aspect ratio holding
# 277 rows, 190 of which carried a single dot. Height is now a stated function of
# the row count, so it is checkable: `24 * n_rows + chrome`, chrome <= 140.
GANTT_ROW_PITCH_PX = 24
GANTT_WIDTH_PX = 1100
GANTT_DPI = 100  # saved pixels == figsize inches * dpi, so the output size is exact
GANTT_LEFT_PX = 340  # row labels: name + record count + stratum label
GANTT_RIGHT_PX = 24
GANTT_TOP_PX = 76  # title + the three declaration lines
GANTT_BOTTOM_PX = 64  # year ticks + axis label + legend
GANTT_CHROME_PX = GANTT_TOP_PX + GANTT_BOTTOM_PX  # 140, the spec's ceiling


def _gantt_geometry(n_rows: int) -> dict[str, Any]:
    """
    Figure size and axes rectangle for `n_rows` rows, in exact pixels.

    Separated from the drawing so the height rule can be asserted without
    rendering anything: the defect being fixed here was a size, and a size is
    only a fix if it is checkable.
    """
    height_px = GANTT_ROW_PITCH_PX * n_rows + GANTT_CHROME_PX
    plot_px = GANTT_ROW_PITCH_PX * n_rows
    return {
        "width_px": GANTT_WIDTH_PX,
        "height_px": height_px,
        "dpi": GANTT_DPI,
        "figsize": (GANTT_WIDTH_PX / GANTT_DPI, height_px / GANTT_DPI),
        "rect": (
            GANTT_LEFT_PX / GANTT_WIDTH_PX,
            GANTT_BOTTOM_PX / height_px,
            (GANTT_WIDTH_PX - GANTT_LEFT_PX - GANTT_RIGHT_PX) / GANTT_WIDTH_PX,
            plot_px / height_px,
        ),
    }


def _gantt_rows(
    records: list[dict],
    pi_name: str = "",
    exclude_names: set[str] | None = None,
) -> dict[str, Any]:
    """
    Select and order the timeline's rows. Visual spec 4.1-4.3.

    Rows are strata A and B: everyone who holds a first-author slot, plus
    everyone who appears twice or more, minus anyone who ever held the
    last-author slot. That is the population the profile report's aggregates are
    already computed over, so the picture and the numbers finally describe the
    same people. Strata C and D are counted and returned rather than drawn, so
    the figure can declare on its own face who is missing from it.

    Note the edge case the stratum rule creates and a plain "appears twice"
    filter would not: one appearance in the first-author slot is stratum A and
    gets a row, while one appearance in the last-author slot is stratum D and
    does not. `_stratum` tests last-author first.

    Ordering is (first appearance date, name), matching `roles.build_people`.
    The previous secondary key `-len(appearances)` ranked people by output
    within a shared first date, which is a ranking of people (spec R5).

    Two limits are inherent to this input and are printed on the chart rather
    than hidden. People are keyed by exact name string, because
    `build_author_records` keeps only `name` — the profile report merges "Smith
    J" and "Smith JA" through ORCID and loose keying, and this renderer cannot.
    And `is_first` here is bare index 0, without the collective-name guard
    `roles.build_people` applies.
    """
    # Lazy: profile.roles imports `_date_iso` from this module, so a top-level
    # import would be circular. Shared rather than copied because a second
    # implementation of the stratum rule is a second thing to drift.
    from pubmed_toolkit.profile.roles import _stratum

    # Case-insensitive, matching `build_people`'s exclusion. A case mismatch
    # otherwise reinstates on the chart someone the report's numbers exclude.
    exclude = {str(name).strip().lower() for name in (exclude_names or ())}
    exclude |= {str(pi_name).strip().lower(), ""}

    people: dict[str, dict[str, Any]] = {}
    corpus_years: list[int] = []
    n_undated = 0

    for rec in records:
        pubdate = str(rec.get("pubdate", ""))
        head = pubdate[:4]
        year = int(head) if head.isdigit() else 0
        # `_date_iso` falls back to 1900 when it finds no year at all. Keeping
        # such a record would stretch the axis over a century of empty columns
        # to place marks on a date nobody published on.
        if year <= 1900:
            n_undated += 1
            continue
        corpus_years.append(year)
        for author in rec.get("authors", []):
            name = str(author.get("name", "")).strip()
            if not name or name.lower() in exclude:
                continue
            person = people.setdefault(name, {
                "name": name,
                "first_date": pubdate,
                "years": set(),
                "lead_years": set(),
                "equal_years": set(),
                "n_appearances": 0,
                "n_first_slots": 0,
                "n_last_slots": 0,
            })
            person["first_date"] = min(person["first_date"], pubdate)
            person["n_appearances"] += 1
            person["years"].add(year)
            if author.get("is_first"):
                person["n_first_slots"] += 1
                person["lead_years"].add(year)
            if author.get("is_last"):
                person["n_last_slots"] += 1
            if author.get("equal_contrib"):
                person["equal_years"].add(year)

    window_start = min(corpus_years) if corpus_years else 0
    window_end = max(corpus_years) if corpus_years else 0

    rows: list[dict[str, Any]] = []
    omitted: Counter[str] = Counter()
    for person in people.values():
        stratum = _stratum(person["n_last_slots"], person["n_first_slots"], person["n_appearances"])
        if stratum not in ("A", "B"):
            omitted[stratum] += 1
            continue
        first_year, last_year = min(person["years"]), max(person["years"])
        rows.append({
            "name": person["name"],
            "stratum": stratum,
            "first_date": person["first_date"],
            "n_appearances": person["n_appearances"],
            "years": sorted(person["years"]),
            "lead_years": sorted(person["lead_years"]),
            "equal_years": sorted(person["equal_years"]),
            "first_year": first_year,
            "last_year": last_year,
            # The window is the corpus itself: this renderer has no search
            # window to read. Same comparison `build_people` makes, so someone
            # publishing at the edge of the record is marked as still going
            # rather than as departed.
            "left_censored": first_year <= window_start,
            "right_censored": last_year >= window_end - 1,
        })

    rows.sort(key=lambda row: (row["first_date"], row["name"]))
    return {
        "rows": rows,
        "omitted": {"C": omitted["C"], "D": omitted["D"]},
        "n_people": len(people),
        "n_undated_records": n_undated,
        "window": (window_start, window_end),
    }


def _gantt_face_text(layout: dict[str, Any]) -> list[str]:
    """
    The lines that state, on the chart itself, who is not on it and why.

    A selection rule that lives only in the surrounding prose is gone the moment
    the figure is screenshotted and forwarded — and here the rule removes most
    of the roster. Exactly three lines: the chrome budget in `_gantt_geometry`
    has room for three and no more.
    """
    omitted = layout["omitted"]
    plotted, total = len(layout["rows"]), layout["n_people"]
    not_plotted = (
        f"Not plotted: {omitted['C']} appear once and hold no first- or last-author slot; "
        f"{omitted['D']} hold a last-author slot at least once (senior collaborators, not a peer group). "
        "Every name is in _author_paper_matrix.csv."
    )
    if layout["n_undated_records"]:
        not_plotted += f" {layout['n_undated_records']} records carry no usable year and are on no axis."
    return [
        f"{plotted} of {total} co-authors plotted — everyone who holds a first-author slot, plus "
        "everyone appearing twice or more, minus anyone who ever held the last-author slot.",
        not_plotted,
        "Grouped by exact name string, so one person written two ways counts twice. "
        "The PI and every configured exclude_names entry are never rows.",
    ]


def render_gantt(
    author_records_path: str,
    output_dir: str,
    pi_name: str = "",
    exclude_names: set[str] | None = None,
) -> str | None:
    """
    Draw the recurring-co-author activity timeline. Visual spec 4.1-4.5.

    One row per stratum-A/B person, ordered by first appearance then name, at a
    fixed 24 px pitch; strata C and D are counted on the chart face instead of
    given rows. Marks are per calendar year, not per record, because
    `_date_iso` fabricates `month=1, day=1` whenever PubMed omits the month —
    the old date axis turned that fabrication into visible January clusters.

    `exclude_names` drops co-authors who would otherwise crowd the chart — a
    co-PI, a department head, a standing collaborator. It is a parameter rather
    than a constant because the previous version hardcoded the names of six real
    people from one specific lab, which then shipped in a public repository.
    Configure it per run via `advisor.exclude_names`; see
    profile.roles.default_gantt_exclude_names.

    Returns None when no one qualifies for a row: an axis drawn over zero rows
    states "nobody recurs here" in the visual grammar of a chart, and the caller
    should say it in prose instead.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D

    # Single-sourced so the chart and the report's roster table cannot disagree
    # about what a stratum is called. Lazy for the same cycle as in _gantt_rows.
    from pubmed_toolkit.profile.report import STRATUM_LABEL

    with open(author_records_path, encoding="utf-8") as f:
        records = json.load(f)

    layout = _gantt_rows(records, pi_name, exclude_names)
    rows = layout["rows"]
    if not rows:
        return None

    geom = _gantt_geometry(len(rows))
    window_start, window_end = layout["window"]
    height_px = geom["height_px"]

    fig = plt.figure(figsize=geom["figsize"], dpi=geom["dpi"])
    ax = fig.add_axes(geom["rect"])

    colors = {"first": "#d62728", "cofirst": "#ff7f0e", "mid": "#1f77b4"}
    marks: dict[str, list[tuple[int, int]]] = {"first": [], "cofirst": [], "mid": []}
    spans: list[tuple[int, int, int]] = []
    tails: dict[str, list[tuple[int, float, int]]] = {"left": [], "right": []}

    for idx, row in enumerate(rows):
        for year in row["years"]:
            if year in row["lead_years"]:
                marks["first"].append((year, idx))
            elif year in row["equal_years"]:
                marks["cofirst"].append((year, idx))
            else:
                marks["mid"].append((year, idx))
        if row["last_year"] > row["first_year"]:
            spans.append((idx, row["first_year"], row["last_year"]))
        if row["left_censored"]:
            tails["left"].append((row["first_year"], window_start - 0.75, idx))
        if row["right_censored"]:
            tails["right"].append((row["last_year"], window_end + 0.75, idx))

    if spans:
        ax.hlines([s[0] for s in spans], [s[1] for s in spans], [s[2] for s in spans],
                  color="#bbbbbb", lw=1.4, zorder=1)
    # Dashed tail plus a hollow arrowhead, so a censored observation differs from
    # a completed one in shape and not only in colour: someone still publishing
    # at the edge of the record has not been shown to have left.
    for key, marker in (("left", "<"), ("right", ">")):
        items = tails[key]
        if not items:
            continue
        ax.hlines([i[2] for i in items], [i[0] for i in items], [i[1] for i in items],
                  color="#8c8c8c", lw=1.0, linestyles="dashed", zorder=2)
        ax.scatter([i[1] for i in items], [i[2] for i in items], marker=marker,
                   facecolors="none", edgecolors="#8c8c8c", s=34, linewidth=0.9, zorder=3)
    for key, marker, size in (("mid", "o", 34), ("cofirst", "D", 40), ("first", "*", 110)):
        points = marks[key]
        if points:
            ax.scatter([p[0] for p in points], [p[1] for p in points], marker=marker,
                       color=colors[key], s=size, zorder=4, edgecolor="white", linewidth=0.5)

    ax.set_yticks(range(len(rows)))
    ax.set_yticklabels(
        [f"{row['name']} — n={row['n_appearances']} records — {STRATUM_LABEL[row['stratum']]}"
         for row in rows],
        fontsize=7,
    )
    # Inverted rather than sorted downward, so the earliest arrival is at the top
    # and the arrival cascade reads left-to-right, top-to-bottom.
    ax.set_ylim(len(rows) - 0.5, -0.5)
    ax.tick_params(axis="y", length=0)

    span_years = window_end - window_start + 1
    # Thin the labels on long corpora; the ticks stay integer years either way.
    step = max(1, -(-span_years // 28))
    ax.set_xticks(range(window_start, window_end + 1, step))
    ax.set_xlim(window_start - 1.1, window_end + 1.1)
    ax.tick_params(axis="x", labelsize=7)
    ax.grid(True, axis="x", alpha=0.25, ls="--", lw=0.6)
    ax.set_xlabel(
        "Publication year — integer only: the month in `pubdate` is fabricated when PubMed omits it",
        fontsize=7,
    )

    fig.text(6 / GANTT_WIDTH_PX, 1 - 16 / height_px,
             f"{pi_name or 'PI'} lab — recurring co-authors, observed publication years",
             fontsize=10, fontweight="bold", va="top", ha="left")
    fig.text(6 / GANTT_WIDTH_PX, 1 - 32 / height_px, "\n".join(_gantt_face_text(layout)),
             fontsize=6.5, va="top", ha="left", color="#333333", linespacing=1.4)

    # In the chrome, not inside the axes. A legend box floating over the plot
    # hides whichever rows sit under it, and at `loc="lower right"` those are the
    # most recent marks of the most recently arrived people.
    fig.legend(handles=[
        Line2D([0], [0], marker="*", color="w", markerfacecolor=colors["first"], markersize=9,
               label="year with a first-author record"),
        # Not "co-first": `equal_contrib` is position-blind, so shared *senior*
        # authorship was being labelled as shared first authorship.
        Line2D([0], [0], marker="D", color="w", markerfacecolor=colors["cofirst"], markersize=5,
               label="year with an equal-contribution flag (position-blind)"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor=colors["mid"], markersize=5,
               label="year with other records"),
        Line2D([0], [0], color="#bbbbbb", lw=1.4,
               label="first-to-last interval, not presence in the lab"),
        Line2D([0], [0], color="#8c8c8c", lw=1.0, ls="--", marker=">", markerfacecolor="none",
               markeredgecolor="#8c8c8c", markersize=5,
               label="record continues to the window edge (censored)"),
    ], loc="lower center", bbox_to_anchor=(0.5, 2 / height_px), ncol=5, fontsize=6, frameon=False)

    out = os.path.join(output_dir, "student_activity_gantt.png")
    # No bbox_inches="tight": the saved pixel size must equal the computed
    # geometry, otherwise the height rule is unverifiable.
    fig.savefig(out, dpi=geom["dpi"])
    plt.close(fig)
    return out


def _load_topic_records(output_dir: str) -> list[dict]:
    json_path = os.path.join(output_dir, "_topic_extraction.json")
    csv_path = os.path.join(output_dir, "_topic_extraction.csv")
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            return json.load(f)
    if os.path.exists(csv_path):
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    return []


def render_topic_charts(output_dir: str) -> list[str]:
    recs = _load_topic_records(output_dir)
    if not recs:
        return []

    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Noto Sans SC", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    prim = Counter(r.get("primary_bucket", "") for r in recs)
    labels, sizes, colors = [], [], []
    for bucket in ["A", "B", "C", "D", "E", "F"]:
        if prim[bucket] > 0:
            labels.append(f"{BUCKET_NAMES[bucket]}\n({prim[bucket]} 篇)")
            sizes.append(prim[bucket])
            colors.append(BUCKET_COLOR[bucket])

    outputs = []
    if sizes:
        fig, ax = plt.subplots(figsize=(9, 7))
        _, _, autotexts = ax.pie(
            sizes, labels=labels, colors=colors, autopct="%1.0f%%",
            startangle=90, pctdistance=0.78,
            wedgeprops={"edgecolor": "white", "linewidth": 2},
            textprops={"fontsize": 10},
        )
        for text in autotexts:
            text.set_color("white")
            text.set_fontweight("bold")
        ax.set_title("主要研究主题分布\n（每篇按 primary bucket 计 1 票）", fontsize=13, pad=20)
        plt.tight_layout()
        out = os.path.join(output_dir, "topic_pie.png")
        plt.savefig(out, dpi=150, bbox_inches="tight")
        plt.close()
        outputs.append(out)

    years = sorted({int(r["year"]) for r in recs if str(r.get("year", "")).isdigit()})
    if years:
        buckets = ["A", "B", "C", "D", "E", "F"]
        mat = np.zeros((len(buckets), len(years)), dtype=int)
        for rec in recs:
            year_text = str(rec.get("year", ""))
            if not year_text.isdigit():
                continue
            yi = years.index(int(year_text))
            for bucket in [rec.get("primary_bucket", ""), rec.get("secondary_bucket", "")]:
                if bucket in buckets:
                    mat[buckets.index(bucket), yi] += 1
        fig, ax = plt.subplots(figsize=(8, 5))
        im = ax.imshow(mat, cmap="YlOrRd", aspect="auto", vmin=0, vmax=max(1, int(mat.max())))
        ax.set_xticks(range(len(years)))
        ax.set_xticklabels(years)
        ax.set_yticks(range(len(buckets)))
        ax.set_yticklabels([BUCKET_NAMES[b].split("\n")[0] for b in buckets])
        for i in range(mat.shape[0]):
            for j in range(mat.shape[1]):
                value = mat[i, j]
                color = "white" if value >= max(1, int(mat.max())) * 0.6 else "black"
                ax.text(j, i, str(value), ha="center", va="center", color=color, fontsize=12, fontweight="bold")
        cbar = plt.colorbar(im, ax=ax, fraction=0.04, pad=0.04)
        cbar.set_label("该年涉及该主题的论文数（含 secondary）")
        ax.set_title("主题 × 年份热图", fontsize=12)
        ax.set_xlabel("Publication year")
        plt.tight_layout()
        out = os.path.join(output_dir, "topic_year_heatmap.png")
        plt.savefig(out, dpi=150, bbox_inches="tight")
        plt.close()
        outputs.append(out)

    return outputs


def run_analysis(
    output_dir: str = "pubmed_results",
    input_excel: str | None = None,
    pdf_dir: str | None = None,
    papers_json: str | None = None,
    pi_name: str = "",
    exclude_names: set[str] | None = None,
) -> dict[str, Any]:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    excel_path = input_excel or find_latest_excel(output_dir)
    pdf_dir = pdf_dir or os.path.join(output_dir, "pdfs")
    json_path = papers_json or find_latest_json(output_dir)

    papers = read_papers_from_excel(excel_path)
    detailed_papers = load_papers_json(json_path)
    results: dict[str, Any] = {
        "excel": excel_path,
        "papers_json": json_path,
        "papers": len(papers),
    }

    results["paper_corpus"] = build_corpus(papers, output_dir, pdf_dir)
    author_records = build_author_records(papers, detailed_papers, output_dir)
    results["authors_parsed"] = author_records
    results["author_matrix"] = write_author_matrix(author_records, output_dir)
    results["gantt"] = render_gantt(author_records, output_dir, pi_name, exclude_names)
    results["topic_charts"] = render_topic_charts(output_dir)
    if not results["topic_charts"]:
        results["topic_note"] = "No _topic_extraction.json/csv found; topic charts skipped."
    return results
