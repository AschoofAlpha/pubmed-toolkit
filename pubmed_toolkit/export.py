"""Export helpers for paper search results."""

from __future__ import annotations

import csv
import json
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any

HEADERS = [
    "序号", "PMID", "标题", "作者", "角色", "期刊", "发表日期",
    "卷", "期", "页码", "DOI", "PMC ID", "PDF状态", "摘要",
]


def _paper_row(idx: int, paper: dict) -> list:
    return [
        idx,
        paper.get("pmid", ""),
        paper.get("title", ""),
        paper.get("authors_str", ""),
        paper.get("role", ""),
        paper.get("journal", ""),
        paper.get("pub_date", ""),
        paper.get("volume", ""),
        paper.get("issue", ""),
        paper.get("pages", ""),
        paper.get("doi", ""),
        paper.get("pmc_id", ""),
        paper.get("pdf_status", ""),
        (paper.get("abstract") or "")[:500],
    ]


def save_to_excel(papers: Iterable[dict], filepath: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "PubMed论文清单"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, paper in enumerate(papers, 1):
        for col, value in enumerate(_paper_row(idx, paper), 1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in [3, 4, 14]))

    widths = [6, 12, 50, 35, 18, 30, 14, 6, 6, 10, 30, 14, 18, 60]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    wb.save(filepath)


def save_to_csv(papers: Iterable[dict], filepath: str) -> None:
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for idx, paper in enumerate(papers, 1):
            writer.writerow(_paper_row(idx, paper))


def save_to_json(
    papers: Iterable[dict],
    filepath: str,
    provenance: dict | None = None,
) -> None:
    """
    Write the paper list, wrapped in an envelope carrying the search parameters.

    Without `provenance` the file is a bare list, which is what earlier versions
    wrote. That format cannot say how many records the search matched, so a
    downstream reader has no way to tell a complete corpus from one silently cut
    off at `retmax` — it can only report the truncation gate as `unknown`, which
    is a gate that never fires. The envelope closes that hole; `load_papers_json`
    still reads the bare list so existing files keep working.
    """
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    payload: Any = list(papers)
    if provenance is not None:
        payload = {
            "schema_version": 1,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "search": provenance,
            "papers": payload,
        }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_papers_json(filepath: str) -> tuple[list[dict], dict]:
    """
    Read either format. Returns (papers, search_provenance).

    A bare list yields an empty provenance dict rather than invented values: the
    caller must be able to tell "this file predates provenance" from "the search
    matched nothing", and a fabricated zero would erase that distinction.
    """
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data, {}
    return data.get("papers") or [], data.get("search") or {}
